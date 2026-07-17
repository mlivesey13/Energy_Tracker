
import sys
print("AUTOMATOR PYTHON:", sys.executable)


import os
import re
import pdfplumber
from pathlib import Path
from openpyxl import Workbook, load_workbook
from datetime import datetime
import matplotlib.pyplot as plt

BILLS_DIR = Path("/Users/marclivesey/Energy_Bills")
EXCEL_PATH = BILLS_DIR / "Energy_Tracker.xlsx"

# Helpers

def ensure_graphs_sheet():
    wb = load_workbook(EXCEL_PATH)
    if "Graphs" not in wb.sheetnames:
        wb.create_sheet("Graphs")
        wb.save(EXCEL_PATH)


# Regex patterns for EDF bill fields
PATTERNS = {
    "balance_last_bill": r"Balance on your last bill\s+£?([\d\.]+)",

    "electricity_used": r"Electricity used\s+([\d\.]+)\s*kWh",
    "electricity_total": r"Total electricity charges for this period\s+£?([\d\.]+)",

    "gas_used": r"Energy used\*?\s+([\d\.]+)\s*kWh",
    "gas_total": r"Total gas charges for this period\s+£?([\d\.]+)",

    # Shared pattern for both electricity & gas subtotals
    "charges_subtotal": r"Subtotal of charges before VAT\s+£?([\d\.]+)",

    "new_balance": r"Your new balance\s+£?([\d\.]+)"
}


def extract_text_from_pdf(pdf_path):
    """Extract full text from a PDF file."""
    with pdfplumber.open(pdf_path) as pdf:
        return "\n".join(page.extract_text() or "" for page in pdf.pages)

def extract_payments(text):
    """Extract all payments (DD + ad-hoc) and return the total."""
    pattern = r"(?:payment|refund)[^\£]*£\s*([\d\.\-]+)"
    matches = re.findall(pattern, text, re.IGNORECASE)
    return sum(float(m) for m in matches) if matches else 0.0

def extract_fields(text):
    """Extract all required fields using regex, including two subtotals."""
    data = {}

    # --- Handle the shared subtotal pattern (electricity first, gas second) ---
    subtotal_matches = re.findall(PATTERNS["charges_subtotal"], text, re.IGNORECASE)

    data["charges_electricity"] = float(subtotal_matches[0]) if len(subtotal_matches) >= 1 else None
    data["charges_gas"] = float(subtotal_matches[1]) if len(subtotal_matches) >= 2 else None

    # --- Handle all other single-match patterns ---
    for key, pattern in PATTERNS.items():
        if key == "charges_subtotal":
            continue  # already handled above

        match = re.search(pattern, text, re.IGNORECASE)
        data[key] = float(match.group(1)) if match else None

    return data


def ensure_excel_exists():
    """Create Excel file with headers if it doesn't exist."""
    if not EXCEL_PATH.exists():
        wb = Workbook()

        # Main sheet
        ws = wb.active
        ws.title = "Energy Data"
        ws.append([
            "File", "Month", "Year", "Balance Last Bill", "Electricity Charges",
            "Gas Charges", "Payments", "New Balance", "Electricity Used (kWh)",
            "Electricity Total (£)", "Gas Used (kWh)", "Gas Total (£)"
        ])

        # Summary sheet
        summary = wb.create_sheet("Summary")
        summary.append([
            "Month", "Year", "Period", "Elec Used (kWh)", "Elec Cost (£)",
            "Gas Used (kWh)", "Gas Cost (£)"
        ])


        wb.save(EXCEL_PATH)

def get_logged_files():
    """Return a set of filenames already logged in Excel."""
    wb = load_workbook(EXCEL_PATH)
    ws = wb["Energy Data"]
    return {row[0].value for row in ws.iter_rows(min_row=2)}

def append_to_excel(filename, month, year, data, text):
    """Append extracted data to Excel workbook."""
    wb = load_workbook(EXCEL_PATH)
    ws = wb["Energy Data"]

    payments_total = extract_payments(text)

    ws.append([
        filename, month, year,
        data["balance_last_bill"],
        data["charges_electricity"],
        data["charges_gas"],
        payments_total,
        data["new_balance"],
        data["electricity_used"],
        data["electricity_total"],
        data["gas_used"],
        data["gas_total"]
    ])

    wb.save(EXCEL_PATH)

def update_summary_sheet():
    """Rebuild the summary sheet from the main data without removing formatting."""
    wb = load_workbook(EXCEL_PATH)
    ws = wb["Energy Data"]
    summary = wb["Summary"]

    # Clear only cell VALUES (not formatting)
    for row in summary.iter_rows(min_row=2):
        for cell in row:
            cell.value = None

    # Write new summary data
    write_row = 2
    for row in ws.iter_rows(min_row=2, values_only=True):
        filename, month, year, _, _, _, _, _, elec_used, elec_cost, gas_used, gas_cost = row

        period = f"{month} {year}"

        summary.cell(row=write_row, column=1, value=month)
        summary.cell(row=write_row, column=2, value=year)
        summary.cell(row=write_row, column=3, value=period)
        summary.cell(row=write_row, column=4, value=elec_used)
        summary.cell(row=write_row, column=5, value=elec_cost)
        summary.cell(row=write_row, column=6, value=gas_used)
        summary.cell(row=write_row, column=7, value=gas_cost)

        write_row += 1

    print("\n➡ Sorting Summary sheet…")
    sort_summary_sheet()
    print("✔ Summary sheet sorted")

    wb.save(EXCEL_PATH)

def sort_summary_sheet():
    """Sort the Summary sheet by Year then Month without removing formatting."""
    wb = load_workbook(EXCEL_PATH)
    summary = wb["Summary"]

    month_order = {
        "January": 1, "February": 2, "March": 3, "April": 4,
        "May": 5, "June": 6, "July": 7, "August": 8,
        "September": 9, "October": 10, "November": 11, "December": 12
    }

    # Load rows (skip header)
    rows = list(summary.iter_rows(min_row=2, values_only=True))

    # Sort by year then month order
    rows_sorted = sorted(
        rows,
        key=lambda r: (r[1], month_order.get(r[0], 0))
    )

    # Overwrite values only (preserve formatting)
    row_index = 2
    for row in rows_sorted:
        for col_index, value in enumerate(row, start=1):
            summary.cell(row=row_index, column=col_index, value=value)
        row_index += 1

    wb.save(EXCEL_PATH)


def detect_missing_months():
    """Detect missing months starting from May 2024 up to the current month."""
    wb = load_workbook(EXCEL_PATH)
    ws = wb["Energy Data"]

    # Extract (month, year) pairs already logged
    months_present = {(row[1].value, row[2].value) for row in ws.iter_rows(min_row=2)}

    # Month ordering
    month_names = [
        "January", "February", "March", "April", "May", "June",
        "July", "August", "September", "October", "November", "December"
    ]

    start_year = 2024
    start_month_index = month_names.index("May")  # May 2024 start

    now = datetime.now()
    current_year = now.year
    current_month_index = now.month - 1  # 0-based index

    # Build list of required months from May 2024 → current month
    required_months = []
    for year in range(start_year, current_year + 1):
        # Determine month range for each year
        if year == start_year:
            month_range = range(start_month_index, 12)
        elif year == current_year:
            month_range = range(0, current_month_index + 1)
        else:
            month_range = range(0, 12)

        for m in month_range:
            required_months.append((month_names[m], year))

    # Determine missing months
    missing = [m for m in required_months if m not in months_present]

    print("\n🔍 Missing Bills (May 2024 → Present):")
    if not missing:
        print("✔ No missing bills detected")
    else:
        for m in missing:
            print(f"- {m[0]} {m[1]}")


def load_summary_data():
    """Load summary sheet data into lists for plotting."""
    wb = load_workbook(EXCEL_PATH)
    ws = wb["Summary"]

    months = []
    elec_used = []
    elec_cost = []
    gas_used = []
    gas_cost = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        month, year, period, eu, ec, gu, gc = row

        months.append(period)   # Use the new Period column
        elec_used.append(eu)
        elec_cost.append(ec)
        gas_used.append(gu)
        gas_cost.append(gc)

    return months, elec_used, elec_cost, gas_used, gas_cost



def plot_graph(x, y, title, ylabel, filename):
    """Generic graph plotting function."""
    plt.figure(figsize=(10, 4))
    plt.plot(x, y, marker='o', linestyle='-', color='blue')
    plt.title(title)
    plt.xlabel("Month")
    plt.ylabel(ylabel)
    plt.xticks(rotation=45, ha='right')
    plt.grid(True)
    plt.tight_layout()
    plt.savefig(BILLS_DIR / filename)
    plt.close()


def generate_usage_and_cost_graphs():
    months, elec_used, elec_cost, gas_used, gas_cost = load_summary_data()

    # Electricity usage
    elec_usage_file = BILLS_DIR / "electricity_usage.png"
    plot_graph(months, elec_used, "Electricity Usage Over Time",
               "Electricity Used (kWh)", elec_usage_file.name)
    insert_graph_into_excel(elec_usage_file, "B2")

    # Gas usage
    gas_usage_file = BILLS_DIR / "gas_usage.png"
    plot_graph(months, gas_used, "Gas Usage Over Time",
               "Gas Used (kWh)", gas_usage_file.name)
    insert_graph_into_excel(gas_usage_file, "B30")

    print("\n📊 Usage and cost graphs generated and added to Excel!")



def sort_energy_data_sheet():
    """Sort the Energy Data sheet by Year then Month without removing formatting."""
    wb = load_workbook(EXCEL_PATH)
    ws = wb["Energy Data"]

    month_order = {
        "January": 1, "February": 2, "March": 3, "April": 4,
        "May": 5, "June": 6, "July": 7, "August": 8,
        "September": 9, "October": 10, "November": 11, "December": 12
    }

    rows = list(ws.iter_rows(min_row=2, values_only=True))

    rows_sorted = sorted(
        rows,
        key=lambda r: (r[2], month_order.get(r[1], 0))
    )

    # Overwrite values only
    row_index = 2
    for row in rows_sorted:
        for col_index, value in enumerate(row, start=1):
            ws.cell(row=row_index, column=col_index, value=value)
        row_index += 1

    wb.save(EXCEL_PATH)

from openpyxl.drawing.image import Image as XLImage

def insert_graph_into_excel(image_path, cell="B2"):
    wb = load_workbook(EXCEL_PATH)
    graphs = wb["Graphs"]

    img = XLImage(image_path)

    # Resize image to fit Excel sheet nicely
    img.width = 900   # pixels
    img.height = 400  # pixels

    graphs.add_image(img, cell)
    wb.save(EXCEL_PATH)



if __name__ == "__main__":
    print("➡ Starting Energy Tracker…")
    print(f"➡ Using Excel file: {EXCEL_PATH}")

    ensure_excel_exists()
    ensure_graphs_sheet()

    print("✔ Excel file checked/created")

    logged = get_logged_files()
    print(f"➡ Logged files found: {len(logged)}")

    pdf_files = list(BILLS_DIR.glob("*.pdf"))
    print(f"➡ PDFs found: {len(pdf_files)}")

    for pdf in pdf_files:
        print(f"\n➡ Processing {pdf.name}…")

        if pdf.name in logged:
            print(f"   ↳ Skipping (already logged)")
            continue

        text = extract_text_from_pdf(pdf)
        data = extract_fields(text)

        parts = pdf.stem.split("_")
        month = parts[1]
        year = int(parts[2])

        append_to_excel(pdf.name, month, year, data, text)
        print(f"   ✔ Added {pdf.name} to Excel")

    print("\n➡ Sorting Energy Data sheet…")
    sort_energy_data_sheet()
    print("✔ Energy Data sorted by date")

    print("\n➡ Updating summary sheet…")
    update_summary_sheet()
    print("✔ Summary sheet updated")

    print("\n➡ Detecting missing months…")
    detect_missing_months()

    print("\n➡ Generating graphs…")
    generate_usage_and_cost_graphs()

    print("\n🎉 ALL TASKS COMPLETED SUCCESSFULLY")
